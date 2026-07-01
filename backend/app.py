import json
import re
import sys
import io
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from flask import Flask, jsonify, request, send_file, Response, send_from_directory
from flask_cors import CORS
import importlib.metadata
import werkzeug
from werkzeug.utils import secure_filename
from werkzeug.exceptions import HTTPException

try:
    werkzeug.__version__
except AttributeError:
    try:
        werkzeug.__version__ = importlib.metadata.version('werkzeug')
    except Exception:
        werkzeug.__version__ = '3.1.8'

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

CORS(
    app,
    resources={r"/*": {"origins": "*"}},
    allow_headers=["Content-Type"],
    methods=["GET", "POST", "OPTIONS"],
    supports_credentials=False,
)

BASE_DIR = Path(__file__).resolve().parent
RULES_PATH = BASE_DIR / 'rules.json'
UPLOAD_DIR = BASE_DIR / 'uploads'
GENERATED_DIR = BASE_DIR / 'generated'
LOG_DIR = BASE_DIR / 'logs'

app.config['UPLOAD_FOLDER'] = str(UPLOAD_DIR)
app.config['GENERATED_FOLDER'] = str(GENERATED_DIR)

# Global cache for validated files (In-memory storage)
VALIDATED_FILES_CACHE = {}


def ensure_runtime_directories() -> None:
    for directory in (UPLOAD_DIR, GENERATED_DIR, LOG_DIR):
        directory.mkdir(parents=True, exist_ok=True)


ensure_runtime_directories()

@app.errorhandler(Exception)
def handle_unexpected_error(error: Exception):
    if isinstance(error, HTTPException):
        return jsonify({"success": False, "message": error.description}), error.code
    print(f"[ERROR] Unhandled exception: {error}")
    return jsonify({"success": False, "message": "Excel processing failed"}), 500


@app.route('/', methods=['GET'])
def root():
    return jsonify({"status": "ok", "message": "BAHL Validation System backend is running"})

@app.before_request
def log_request_context():
    if request.path in {"/analyze", "/validate"}:
        print(f"[REQUEST] {request.method} {request.path}")

def load_rules() -> Dict[str, Any]:
    try:
        with RULES_PATH.open('r', encoding='utf-8') as fh:
            return json.load(fh)
    except FileNotFoundError:
        print('[STARTUP] rules.json not found; continuing with empty rules')
        return {'profiles': {}}
    except json.JSONDecodeError as exc:
        print(f'[STARTUP] Invalid rules.json: {exc}')
        return {'profiles': {}}


RULES = load_rules()

SUPPORTED_EXTENSIONS = {'.xlsx', '.xls'}
XLSX_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml'
}
XLS_MIME_TYPES = {'application/vnd.ms-excel', 'application/xls'}

XLSX_TEXT_TYPES = {'s', 'str', 'inlineStr'}
XLSX_NUMERIC_TYPES = {'n'}
XLSX_DATE_TYPES = {'d'}
XLSX_BOOLEAN_TYPES = {'b'}
XLSX_FORMULA_TYPES = {'f'}
XLSX_ERROR_TYPES = {'e'}

TEXT_ONLY_FIELDS = {
    'BeneficiaryName',
    'Amount',
    'BeneficiaryMobile',
    'BeneficiaryNumber',
    'BeneficiaryIdentificationType',
    'BeneficiaryIdentificationNo',
    'BeneficiaryIBAN',
    'BeneficiaryAccountNumber',
    'BeneficiaryAccountNo',
    'ProductTypeCode'
}


def detect_excel_format(filename: Optional[str], mime_type: Optional[str], file_bytes: Optional[bytes]) -> Dict[str, Any]:
    if not file_bytes:
        return {'kind': 'unsupported', 'reason': 'Uploaded file is empty'}

    if not isinstance(file_bytes, (bytes, bytearray, memoryview)):
        file_bytes = bytes(file_bytes)

    filename = (filename or '').lower()
    extension = Path(filename).suffix.lower()
    mime = (mime_type or '').lower()
    header = file_bytes[:8]
    has_zip_signature = header.startswith(b'PK\x03\x04')
    has_ole_signature = header.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1')

    unsupported_extensions = {'.csv', '.xlsm', '.xltx', '.xltm', '.txt', '.invalid'}
    if extension in unsupported_extensions:
        return {'kind': 'unsupported', 'reason': 'Unsupported Excel format'}

    if extension == '.xlsx':
        if mime in XLSX_MIME_TYPES or mime in {'application/zip', 'application/octet-stream', ''} or has_zip_signature:
            return {'kind': 'xlsx', 'reason': None}
        return {'kind': 'unsupported', 'reason': 'Unsupported Excel format'}

    if extension == '.xls':
        if mime in XLS_MIME_TYPES or mime in {'application/octet-stream', ''} or has_ole_signature:
            return {'kind': 'xls', 'reason': None}
        return {'kind': 'unsupported', 'reason': 'Unsupported Excel format'}

    if has_zip_signature and mime in XLSX_MIME_TYPES or mime in {'application/zip', 'application/octet-stream', ''}:
        return {'kind': 'xlsx', 'reason': None}

    if has_ole_signature and mime in XLS_MIME_TYPES or mime in {'application/octet-stream', ''}:
        return {'kind': 'xls', 'reason': None}

    return {'kind': 'unsupported', 'reason': 'Unsupported Excel format'}


def _to_cell_text(value: Any, excel_type: str = 'text') -> str:
    if value is None:
        return ''
    if excel_type == 'text' and isinstance(value, str):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip('0').rstrip('.') if '.' in str(value) else str(value)
    if isinstance(value, bool):
        return str(value)
    return str(value)


def _openpyxl_cell_type_label(cell) -> str:
    dtype = getattr(cell, 'data_type', None)
    if dtype in XLSX_TEXT_TYPES:
        return 'text'
    if dtype in XLSX_NUMERIC_TYPES:
        return 'number'
    if dtype in XLSX_DATE_TYPES:
        return 'date'
    if dtype in XLSX_BOOLEAN_TYPES:
        return 'boolean'
    if dtype in XLSX_FORMULA_TYPES:
        return 'formula'
    if dtype in XLSX_ERROR_TYPES:
        return 'error'
    if isinstance(cell.value, str):
        return 'text'
    if isinstance(cell.value, (int, float)):
        return 'number'
    return 'text'


def _xlrd_cell_type_label(cell_type: int) -> str:
    try:
        import xlrd
        if cell_type == xlrd.XL_CELL_TEXT:
            return 'text'
        if cell_type == xlrd.XL_CELL_NUMBER:
            return 'number'
        if cell_type == xlrd.XL_CELL_DATE:
            return 'date'
        if cell_type == xlrd.XL_CELL_BOOLEAN:
            return 'boolean'
        if cell_type == xlrd.XL_CELL_ERROR:
            return 'error'
        return 'text'
    except Exception:
        return 'text'


def _clean_text_for_validation(value: str) -> str:
    return value.strip() if isinstance(value, str) else _to_cell_text(value)


def _get_row_cell_meta(row: Dict[str, Any], field: str) -> Dict[str, Any]:
    return row.get('_meta', {}).get(field, {})


def _has_hidden_whitespace(value: str) -> bool:
    if not isinstance(value, str) or value == '':
        return False
    return bool(re.search(r'[\u00A0\u200B\u200C\u200D\uFEFF]', value))


def _has_multiple_spaces(value: str) -> bool:
    if not isinstance(value, str):
        return False
    return bool(re.search(r' {2,}', value))


def parse_excel_workbook(file_bytes: bytes, filename: Optional[str], mime_type: Optional[str]) -> tuple[List[str], List[Dict[str, str]], str]:
    detection = detect_excel_format(filename, mime_type, file_bytes)
    if detection['kind'] == 'unsupported':
        raise ValueError(detection['reason'])

    print(f"[UPLOAD] Detected type: {detection['kind']}")

    if detection['kind'] == 'xlsx':
        print('[PARSER] Parser selected: openpyxl')
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            worksheet = workbook.active
            sheet_name = worksheet.title or 'Sheet1'
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            headers: List[str] = []
            rows: List[Dict[str, str]] = []

            for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                if row_index == 1:
                    headers = [_to_cell_text(cell.value, _openpyxl_cell_type_label(cell)) for cell in row]
                    continue

                row_data: Dict[str, Any] = {}
                row_meta: Dict[str, Any] = {}
                has_value = False

                for col_idx, cell in enumerate(row):
                    header = headers[col_idx] if col_idx < len(headers) else f'Column{col_idx + 1}'
                    cell_type = _openpyxl_cell_type_label(cell)
                    cell_text = _to_cell_text(cell.value, cell_type)
                    if cell_text != '':
                        has_value = True
                    row_data[header or f'Column{col_idx + 1}'] = cell_text
                    row_meta[header or f'Column{col_idx + 1}'] = {
                        'excel_type': cell_type,
                        'raw_value': cell.value,
                        'text': cell_text
                    }

                if not has_value:
                    continue

                row_data['_meta'] = row_meta
                rows.append(row_data)

            workbook.close()
            print(f"[UPLOAD] Rows extracted: {len(rows)}")
            return headers, rows, sheet_name
        except Exception as exc:
            print(f"[PARSER] XLSX parse failed: {exc}")
            raise ValueError('Uploaded Excel file is corrupted.') from exc

    print('[PARSER] Parser selected: xlrd')
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=file_bytes)
        worksheet = workbook.sheet_by_index(0)
        headers: List[str] = []
        rows: List[Dict[str, str]] = []

        for row_index in range(worksheet.nrows):
            if row_index == 0:
                headers = [_to_cell_text(worksheet.cell_value(row_index, col_index), 'text') for col_index in range(worksheet.ncols)]
                continue

            row_data: Dict[str, Any] = {}
            row_meta: Dict[str, Any] = {}
            has_value = False

            for col_idx in range(worksheet.ncols):
                cell = worksheet.cell(row_index, col_idx)
                cell_type = _xlrd_cell_type_label(cell.ctype)
                cell_value = cell.value
                if cell_type == 'date':
                    try:
                        cell_text = str(xlrd.xldate_as_datetime(cell_value, workbook.datemode))
                    except Exception:
                        cell_text = _to_cell_text(cell_value, cell_type)
                else:
                    cell_text = _to_cell_text(cell_value, cell_type)

                if cell_text != '':
                    has_value = True
                header = headers[col_idx] if col_idx < len(headers) else f'Column{col_idx + 1}'
                row_data[header or f'Column{col_idx + 1}'] = cell_text
                row_meta[header or f'Column{col_idx + 1}'] = {
                    'excel_type': cell_type,
                    'raw_value': cell_value,
                    'text': cell_text
                }

            if not has_value:
                continue

            row_data['_meta'] = row_meta
            rows.append(row_data)

        print(f"[UPLOAD] Rows extracted: {len(rows)}")
        return headers, rows, worksheet.name
    except Exception as exc:
        print(f"[PARSER] XLS parse failed: {exc}")
        raise ValueError('Uploaded Excel file is corrupted.') from exc


PROFILE_HEADERS = {
    'raast': [
        'Amount', 'BeneficiaryBankCode', 'BeneficiaryIBAN', 'BeneficiaryName', 'BeneficiaryCode',
        'ReferenceField1', 'ReferenceField2', 'BeneficiaryEmail', 'BeneficiaryMobile',
        'BeneficiaryIdentificationType', 'BeneficiaryIdentificationNo', 'ProductTypeCode'
    ],
    'web_based': [
        'Amount', 'BeneficiaryBankCode', 'BeneficiaryAccountNo', 'BeneficiaryName', 'BeneficiaryCode',
        'ReferenceField1', 'ReferenceField2', 'BeneficiaryEmail', 'BeneficiaryMobile', 'ProductTypeCode'
    ],
    '1link': [
        'Amount', 'BankName', 'BeneficiaryAccountNumber', 'BeneficiaryName', 'BeneficiaryCode',
        'ReferenceField1', 'ReferenceField2', 'BeneficiaryEmail', 'BeneficiaryNumber'
    ],
}

def clean_cell_value(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        val_str = f"{val:.20f}"
        if '.' in val_str:
            val_str = val_str.rstrip('0').rstrip('.')
        return val_str
    if isinstance(val, int):
        return str(val)
    return str(val).strip()

def detect_profile(headers: List[str]) -> str:
    normalized = {h.strip(): h for h in headers if h}
    
    # Check for direct indicators of profile
    if any(h in normalized for h in ['BeneficiaryIBAN', 'BeneficiaryIdentificationType', 'BeneficiaryIdentificationNo']):
        return 'raast'
    if any(h in normalized for h in ['BeneficiaryAccountNo']):
        return 'web_based'
    if any(h in normalized for h in ['BeneficiaryAccountNumber', 'BeneficiaryNumber']):
        return '1link'
        
    # Check exact matching fallback
    for profile_name, expected_headers in PROFILE_HEADERS.items():
        if all(header in normalized for header in expected_headers):
            return profile_name
            
    return 'unknown'


def validate_rows(rows: List[Dict[str, str]], profile: str, rules: Dict[str, Any], debug: bool = False) -> List[Dict[str, Any]]:
    headers = PROFILE_HEADERS.get(profile, [])
    return validate_excel_data(rows, headers, profile, rules, debug=debug)


def validate_excel_data(rows: List[Dict[str, str]], headers: List[str], profile: str, rules: Dict[str, Any], debug: bool = False) -> List[Dict[str, Any]]:
    header_map = {h.strip(): i for i, h in enumerate(headers) if h}
    
    # Precompute duplicates
    duplicate_tracker = {
        "BeneficiaryAccountNumber": {}
    }
    for idx, row in enumerate(rows):
        for field in duplicate_tracker.keys():
            val = row.get(field, '').strip()
            if val:
                duplicate_tracker[field].setdefault(val, []).append(idx)

    validation_results = []
    
    for idx, row in enumerate(rows):
        errors = []
        profile_rules = rules['profiles'][profile]
        mandatory_fields = profile_rules['mandatory']

        row_meta = row.get('_meta', {})

        def report_type_error(field_name, actual_type, msg=None):
            errors.append({
                'col': header_map.get(field_name),
                'field': field_name,
                'msg': msg or 'Cell format is not text',
                'expected': 'Text',
                'actual': actual_type
            })

        def report_space_error(field_name, actual_value, msg=None):
            errors.append({
                'col': header_map.get(field_name),
                'field': field_name,
                'msg': msg or 'Leading/trailing spaces not allowed',
                'expected': 'Text without extra spaces',
                'actual': actual_value
            })

        # 0. Required text-only fields should be stored as text
        for field in TEXT_ONLY_FIELDS:
            if field not in header_map:
                continue
            raw_meta = row_meta.get(field, {})
            excel_type = raw_meta.get('excel_type', 'text')
            actual_value = raw_meta.get('text', row.get(field, ''))
            if actual_value != '':
                if excel_type != 'text':
                    report_type_error(field, excel_type, 'Cell format is not Text')
                if isinstance(actual_value, str):
                    if actual_value != actual_value.strip():
                        report_space_error(field, actual_value, 'Leading/trailing spaces not allowed')
                    elif _has_multiple_spaces(actual_value):
                        report_space_error(field, actual_value, 'Multiple spaces are not allowed')
                    elif _has_hidden_whitespace(actual_value):
                        report_space_error(field, actual_value, 'Hidden whitespace characters are not allowed')

        # 1. Mandatory Fields Check
        for field in mandatory_fields:
            val = row.get(field, '')
            if isinstance(val, str):
                if val.strip() == '':
                    errors.append({
                        'col': header_map.get(field),
                        'field': field,
                        'msg': f'{field} is required',
                        'expected': 'Non-empty value',
                        'actual': '(Empty)'
                    })
            else:
                errors.append({
                    'col': header_map.get(field),
                    'field': field,
                    'msg': f'{field} is required',
                    'expected': 'Text',
                    'actual': type(val).__name__
                })

        # 2. Amount Check
        amount_val_raw = row.get('Amount', '')
        amount_val = amount_val_raw.strip() if isinstance(amount_val_raw, str) else _to_cell_text(amount_val_raw)
        if amount_val:
            try:
                amount_float = float(amount_val)
                if amount_float <= 0:
                    errors.append({
                        'col': header_map.get('Amount'),
                        'field': 'Amount',
                        'msg': 'Amount must be greater than 0',
                        'expected': 'Number > 0',
                        'actual': amount_val
                    })
            except ValueError:
                errors.append({
                    'col': header_map.get('Amount'),
                    'field': 'Amount',
                    'msg': 'Amount must be a numeric value',
                    'expected': 'Numeric value',
                    'actual': amount_val
                })

        # 3. Mobile/Telecom check
        mobile_field = 'BeneficiaryNumber' if profile == '1link' else 'BeneficiaryMobile'
        mobile_val_raw = row.get(mobile_field, '')
        mobile_val = mobile_val_raw.strip() if isinstance(mobile_val_raw, str) else _to_cell_text(mobile_val_raw)
        if mobile_val:
            if not re.fullmatch(r'\+?\d{10,15}', mobile_val):
                errors.append({
                    'col': header_map.get(mobile_field),
                    'field': mobile_field,
                    'msg': 'Mobile number must be 11 digits',
                    'expected': 'Telecom format (11 digits)',
                    'actual': mobile_val
                })

        # 4. Identity validation (Profile A/raast only)
        if profile == 'raast':
            id_type_raw = row.get('BeneficiaryIdentificationType', '')
            id_type = id_type_raw.strip() if isinstance(id_type_raw, str) else _to_cell_text(id_type_raw)
            id_no_raw = row.get('BeneficiaryIdentificationNo', '')
            id_no = id_no_raw.strip() if isinstance(id_no_raw, str) else _to_cell_text(id_no_raw)
            iban_raw = row.get('BeneficiaryIBAN', '')
            iban = iban_raw.strip() if isinstance(iban_raw, str) else _to_cell_text(iban_raw)
            
            if id_type:
                if id_type not in ['CNIC', 'NTN']:
                    errors.append({
                        'col': header_map.get('BeneficiaryIdentificationType'),
                        'field': 'BeneficiaryIdentificationType',
                        'msg': 'Identification Type must be CNIC or NTN',
                        'expected': 'CNIC or NTN',
                        'actual': id_type
                    })
                if id_no:
                    if id_type == 'CNIC':
                        if not id_no.isdigit():
                            errors.append({
                                'col': header_map.get('BeneficiaryIdentificationNo'),
                                'field': 'BeneficiaryIdentificationNo',
                                'msg': 'CNIC must be numeric',
                                'expected': 'Numeric value',
                                'actual': id_no
                            })
                        elif len(id_no) != 13:
                            errors.append({
                                'col': header_map.get('BeneficiaryIdentificationNo'),
                                'field': 'BeneficiaryIdentificationNo',
                                'msg': 'CNIC must be exactly 13 digits',
                                'expected': '13 digits (numeric)',
                                'actual': id_no
                            })
                    elif id_type == 'NTN':
                        if not re.fullmatch(r'[A-Za-z0-9]{8}', id_no):
                            if len(id_no) != 8:
                                errors.append({
                                    'col': header_map.get('BeneficiaryIdentificationNo'),
                                    'field': 'BeneficiaryIdentificationNo',
                                    'msg': 'NTN must be exactly 8 characters',
                                    'expected': '8 chars',
                                    'actual': id_no
                                })
                            else:
                                errors.append({
                                    'col': header_map.get('BeneficiaryIdentificationNo'),
                                    'field': 'BeneficiaryIdentificationNo',
                                    'msg': 'NTN must be alphanumeric',
                                    'expected': 'Alphanumeric',
                                    'actual': id_no
                                })
                else:
                    errors.append({
                        'col': header_map.get('BeneficiaryIdentificationNo'),
                        'field': 'BeneficiaryIdentificationNo',
                        'msg': 'Identification Number is required',
                        'expected': 'ID Number',
                        'actual': '(Empty)'
                    })
            elif id_no:
                errors.append({
                    'col': header_map.get('BeneficiaryIdentificationType'),
                    'field': 'BeneficiaryIdentificationType',
                    'msg': 'Identification Type is required',
                    'expected': 'CNIC or NTN',
                    'actual': '(Empty)'
                })

            # 5. IBAN / Account Number rule
            iban_field = 'BeneficiaryIBAN'
            if iban:
                if iban.startswith('PK'):
                    if len(iban) != 24 or not iban.isalnum() or not iban.isupper():
                        errors.append({
                            'col': header_map.get(iban_field),
                            'field': iban_field,
                            'msg': 'IBAN invalid length or format',
                            'expected': '24-char uppercase IBAN',
                            'actual': iban
                        })
                else:
                    if iban == '':
                        errors.append({
                            'col': header_map.get(iban_field),
                            'field': iban_field,
                            'msg': 'Account Number is required',
                            'expected': 'Non-empty Account Number',
                            'actual': '(Empty)'
                        })

        # 6. Duplicate Checks
        for field in duplicate_tracker.keys():
            val = row.get(field, '').strip()
            if val and len(duplicate_tracker[field][val]) > 1:
                errors.append({
                    'col': header_map.get(field),
                    'field': field,
                    'msg': f'Duplicate {field} detected',
                    'expected': 'Unique value',
                    'actual': val,
                    'is_duplicate': True
                })

        if debug:
            profile_label = {
                'raast': 'Raast',
                'web_based': 'Web-Based',
                '1link': '1Link'
            }.get(profile, profile)
            name = row.get('BeneficiaryName', '').strip() or '(blank)'
            if profile == 'raast':
                identifier = row.get('BeneficiaryIdentificationNo', '').strip() or '(blank)'
                id_type = row.get('BeneficiaryIdentificationType', '').strip() or '(blank)'
                detail = f"Name: {name}, Type: {id_type}, ID: {identifier}"
            elif profile == 'web_based':
                identifier = row.get('BeneficiaryAccountNo', '').strip() or '(blank)'
                detail = f"Name: {name}, Account: {identifier}"
            else:
                identifier = row.get('BeneficiaryAccountNumber', '').strip() or '(blank)'
                detail = f"Name: {name}, Account: {identifier}"

            if errors:
                reason = '; '.join(err['msg'] for err in errors)
                print(f"[DEBUG] Processing Row {idx + 1} [Profile: {profile_label}] -> {detail} -> STATUS: FAIL [Reason: {reason}]")
            else:
                bypass_suffix = ' (Identity check bypassed)' if profile in {'web_based', '1link'} else ''
                print(f"[DEBUG] Processing Row {idx + 1} [Profile: {profile_label}] -> {detail} -> STATUS: PASS{bypass_suffix}")

        validation_results.append({
            'status': 'PASS' if not errors else 'FAIL',
            'errors': errors
        })
        
    return validation_results


def _build_excel_styles() -> Dict[str, Any]:
    return {
        'header_fill': PatternFill(start_color='FF0E8348', end_color='FF0E8348', fill_type='solid'),
        'header_font': Font(color='FFFFFFFF', bold=True, size=12),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'error_fill': PatternFill(start_color='FFFDE2E2', end_color='FFFDE2E2', fill_type='solid'),
        'error_border': Border(left=Side(style='thin', color='FFB91C1C'), right=Side(style='thin', color='FFB91C1C'), top=Side(style='thin', color='FFB91C1C'), bottom=Side(style='thin', color='FFB91C1C')),
        'error_font': Font(color='FF8B0000', bold=True),
        'status_valid_fill': PatternFill(start_color='FF0A6B3B', end_color='FF0A6B3B', fill_type='solid'),
        'status_valid_font': Font(color='FFFFFFFF', bold=True),
        'status_invalid_fill': PatternFill(start_color='FFD32F2F', end_color='FFD32F2F', fill_type='solid'),
        'status_invalid_font': Font(color='FFFFFFFF', bold=True),
        'summary_pass_fill': PatternFill(start_color='FF0A6B3B', end_color='FF0A6B3B', fill_type='solid'),
        'summary_fail_fill': PatternFill(start_color='FFD32F2F', end_color='FFD32F2F', fill_type='solid'),
        'summary_time_fill': PatternFill(start_color='FFE5E7EB', end_color='FFE5E7EB', fill_type='solid'),
        'body_alignment': Alignment(vertical='top', wrap_text=True),
    }


def _apply_sheet_styling(ws, styles: Dict[str, Any]) -> None:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        for cell in row:
            if row_idx == 1:
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
                cell.alignment = styles['header_alignment']
            else:
                cell.alignment = styles['body_alignment']
                if cell.value is None:
                    cell.value = ''

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col) if col else 12
        col_letter = get_column_letter(col[0].column) if col else 'A'
        width = min(max(max_len + 3, 15), 40)
        ws.column_dimensions[col_letter].width = width

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.freeze_panes = 'A2'
        ws.sheet_view.showGridLines = True


def _humanize_error(error: Dict[str, Any]) -> str:
    msg = (error.get('msg') or 'Validation error').strip()
    field = (error.get('field') or '').strip().lower()
    normalized = msg.lower()

    if 'mobile' in normalized or 'beneficiarynumber' in field or 'beneficiarymobile' in field:
        return 'Mobile number format invalid'
    if 'iban' in normalized or 'beneficiaryiban' in field:
        return 'IBAN must contain exactly 24 characters'
    if 'cnic' in normalized or 'beneficiaryidentificationno' in field:
        if 'required' in normalized or 'missing' in normalized:
            return 'CNIC required'
        return 'CNIC format invalid'
    if 'duplicate' in normalized:
        if 'mobile' in normalized or 'beneficiarynumber' in field or 'beneficiarymobile' in field:
            return 'Duplicate Mobile Number'
        return 'Duplicate record'
    if 'required' in normalized or 'missing' in normalized:
        return 'Missing mandatory field'
    if 'numeric' in normalized or 'format' in normalized or 'wrong cell' in normalized:
        return 'Wrong cell format'
    if msg.endswith('.'):
        return msg
    return msg + '.'


def _suggested_fix(error: Dict[str, Any]) -> str:
    msg = (error.get('msg') or '').lower()
    if 'iban' in msg:
        return 'Use 24-char IBAN'
    if 'mobile' in msg or 'beneficiarynumber' in msg or 'beneficiarymobile' in msg:
        return 'Use 03XXXXXXXXX'
    if 'cnic' in msg or 'beneficiaryidentificationno' in msg:
        return 'Provide valid CNIC'
    if 'ntn' in msg:
        return 'Provide valid NTN value'
    if 'duplicate' in msg:
        return 'Remove duplicate'
    if 'required' in msg or 'missing' in msg:
        return 'Fill in required field'
    if 'header' in msg:
        return 'Correct header name'
    if 'spaces' in msg:
        return 'Remove leading/trailing spaces'
    if 'format' in msg or 'numeric' in msg:
        return 'Convert to Text'
    return 'Review and correct the value'


def _build_error_summary(errors: List[Dict[str, Any]]) -> str:
    return ' | '.join(_humanize_error(err) for err in errors) if errors else 'None'


def _build_suggested_fix_summary(errors: List[Dict[str, Any]]) -> str:
    return ' ; '.join(_suggested_fix(err) for err in errors) if errors else 'No action required.'


def generate_full_validation_report(rows, results, headers, profile):
    normalized_headers = list(headers or [])
    if not normalized_headers:
        normalized_headers = ['Column1']

    styles = _build_excel_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ValidatedReport'
    ws.append(normalized_headers + ['Validation Status', 'Error Reason', 'Suggested Fix'])

    for idx, row in enumerate(rows):
        if idx >= len(results):
            break

        result = results[idx]
        row_values = [row.get(header, '') for header in normalized_headers]
        error_summary = _build_error_summary(result.get('errors', []))
        suggested_fix = _build_suggested_fix_summary(result.get('errors', []))

        ws.append(row_values + [result['status'], error_summary, suggested_fix])
        current_row = idx + 2

        if result.get('errors'):
            for error in result['errors']:
                field_name = error.get('field', '')
                col_idx = next((i for i, h in enumerate(normalized_headers) if h == field_name), None)
                if col_idx is not None:
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    cell.fill = styles['error_fill']
                    cell.font = styles['error_font']
                    cell.border = styles['error_border']
                    cell.alignment = styles['body_alignment']
                    comment_text = error.get('msg', 'Validation Error')
                    if error.get('expected'):
                        comment_text += f"\nExpected: {error['expected']}"
                    if error.get('actual'):
                        comment_text += f"\nActual: {error['actual']}"
                    cell.comment = Comment(comment_text, 'Validator')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_cell = row[len(normalized_headers)]
        if status_cell.value == 'PASS':
            status_cell.fill = styles['status_valid_fill']
            status_cell.font = styles['status_valid_font']
        else:
            status_cell.fill = styles['status_invalid_fill']
            status_cell.font = styles['status_invalid_font']
        status_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    _apply_sheet_styling(ws, styles)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_highlighted_validation_report(rows, results, headers, profile):
    """Generate a validation report with all rows, highlighting error cells with red background and comments."""
    normalized_headers = list(headers or [])
    if not normalized_headers:
        normalized_headers = ['Column1']

    # Create the report workbook
    wb_report = openpyxl.Workbook()
    ws_report = wb_report.active
    ws_report.title = 'Validation Report'
    
    # Add headers
    ws_report.append(normalized_headers)
    
    styles = _build_excel_styles()
    
    # Create a mapping from field names to column indices for easy error cell highlighting
    field_to_col = {header: idx + 1 for idx, header in enumerate(normalized_headers)}
    
    # Add all rows with error highlighting
    for idx, row in enumerate(rows):
        if idx >= len(results):
            break
        
        result = results[idx]
        row_data = [row.get(header, '') for header in normalized_headers]
        
        # Add the row
        ws_report.append(row_data)
        current_row = idx + 2  # +2 because row 1 is headers, +1 for 1-based indexing
        
        # Highlight error cells and add comments
        if result['errors']:
            for error in result['errors']:
                field_name = error.get('field', '')
                col_idx = field_to_col.get(field_name)
                
                if col_idx:
                    cell = ws_report.cell(row=current_row, column=col_idx)
                    cell.fill = styles['error_fill']
                    cell.font = styles['error_font']
                    cell.border = styles['error_border']
                    cell.alignment = styles['body_alignment']
                    # Add comment with error message
                    error_msg = error.get('msg', 'Validation Error')
                    if error.get('expected'):
                        error_msg += f"\nExpected: {error['expected']}"
                    if error.get('actual'):
                        error_msg += f"\nActual: {error['actual']}"
                    cell.comment = Comment(error_msg, 'Validator')
    
    _apply_sheet_styling(ws_report, styles)
    
    # Serialize to bytes
    report_stream = io.BytesIO()
    wb_report.save(report_stream)
    report_stream.seek(0)
    
    return report_stream.getvalue()

def generate_validated_excel_streams(rows, results, headers, profile):
    """Generate XLSX output workbooks from normalized rows and validation results.
    
    Both .xlsx and .xls inputs are normalized into the same row structure by the parser,
    so this function always outputs modern .xlsx format regardless of input format.
    """
    normalized_headers = list(headers or [])
    if not normalized_headers:
        normalized_headers = ['Column1']

    styles = _build_excel_styles()

    # Create workbooks for passed and rejected records
    wb_passed = openpyxl.Workbook()
    ws_passed = wb_passed.active
    ws_passed.title = 'Sheet'
    ws_passed.append(normalized_headers + ['Validation Status'])

    wb_rejected = openpyxl.Workbook()
    ws_rejected = wb_rejected.active
    ws_rejected.title = 'Sheet'
    ws_rejected.append(normalized_headers + ['Validation Status', 'Error Reason', 'Suggested Fix'])

    # Build the passed and rejected workbooks row by row
    for idx, row in enumerate(rows):
        if idx >= len(results):
            break
        
        result = results[idx]
        row_values = [row.get(h, '') for h in normalized_headers]

        if result['status'] == 'PASS':
            ws_passed.append(row_values + ['PASS'])
        else:
            errors = result.get('errors', [])
            error_summary = _build_error_summary(errors)
            suggested_fix = _build_suggested_fix_summary(errors)
            ws_rejected.append(row_values + ['FAIL', error_summary, suggested_fix])
            rejected_row_number = ws_rejected.max_row

            for err in errors:
                col_idx = err.get('col')
                if col_idx is not None and col_idx < len(normalized_headers):
                    cell = ws_rejected.cell(row=rejected_row_number, column=col_idx + 1)
                    cell.fill = styles['error_fill']
                    cell.font = styles['error_font']
                    cell.border = styles['error_border']
                    cell.alignment = styles['body_alignment']
                    comment_text = f"{_humanize_error(err)}\nExpected: {err.get('expected', 'See rule')}\nActual: {err.get('actual', 'Invalid value')}"
                    cell.comment = Comment(comment_text, 'Validator')

    # Style passed sheet status column
    for row in ws_passed.iter_rows(min_row=2, max_row=ws_passed.max_row):
        status_cell = row[-1]
        status_cell.fill = styles['status_valid_fill']
        status_cell.font = styles['status_valid_font']
        status_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Style rejected sheet status and error columns
    for row in ws_rejected.iter_rows(min_row=2, max_row=ws_rejected.max_row):
        status_cell = row[len(normalized_headers)]
        status_cell.fill = styles['status_invalid_fill']
        status_cell.font = styles['status_invalid_font']
        status_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx in range(len(normalized_headers) + 1, ws_rejected.max_column + 1):
            cell = row[col_idx - 1]
            cell.alignment = styles['body_alignment']
            if cell.value is None:
                cell.value = ''

    # Format both workbooks
    for ws in [ws_passed, ws_rejected]:
        _apply_sheet_styling(ws, styles)
        if ws.max_row > 1:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = styles['body_alignment']
                    if cell.value is None:
                        cell.value = ''

    # Serialize to bytes
    passed_stream = io.BytesIO()
    wb_passed.save(passed_stream)
    passed_stream.seek(0)

    rejected_stream = io.BytesIO()
    wb_rejected.save(rejected_stream)
    rejected_stream.seek(0)

    return passed_stream.getvalue(), rejected_stream.getvalue()


def generate_rejected_validation_report(rows, results, headers, profile):
    normalized_headers = list(headers or [])
    if not normalized_headers:
        normalized_headers = ['Column1']

    styles = _build_excel_styles()
    wb_rejected = openpyxl.Workbook()
    ws_rejected = wb_rejected.active
    ws_rejected.title = 'RejectedRecords'
    ws_rejected.append(normalized_headers + ['Validation Status', 'Error Reason', 'Suggested Fix'])

    for idx, row in enumerate(rows):
        if idx >= len(results):
            break

        result = results[idx]
        if result.get('status') == 'PASS':
            continue

        errors = result.get('errors', [])
        error_summary = _build_error_summary(errors)
        suggested_fix = _build_suggested_fix_summary(errors)
        row_values = [row.get(h, '') for h in normalized_headers]

        ws_rejected.append(row_values + ['FAIL', error_summary, suggested_fix])
        rejected_row_number = ws_rejected.max_row

        for err in errors:
            col_idx = err.get('col')
            if col_idx is not None and 0 <= col_idx < len(normalized_headers):
                cell = ws_rejected.cell(row=rejected_row_number, column=col_idx + 1)
                cell.fill = styles['error_fill']
                cell.font = styles['error_font']
                cell.border = styles['error_border']
                cell.alignment = styles['body_alignment']
                comment_text = err.get('msg', 'Validation Error')
                if err.get('expected'):
                    comment_text += f"\nExpected: {err['expected']}"
                if err.get('actual'):
                    comment_text += f"\nActual: {err['actual']}"
                cell.comment = Comment(comment_text, 'Validator')

    if ws_rejected.max_row > 1:
        for row in ws_rejected.iter_rows(min_row=2, max_row=ws_rejected.max_row):
            status_cell = row[len(normalized_headers)]
            status_cell.fill = styles['status_invalid_fill']
            status_cell.font = styles['status_invalid_font']
            status_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in row:
                if cell.value is None:
                    cell.value = ''
                if cell.fill == styles['header_fill']:
                    continue
                if cell.row == 1:
                    continue
                if cell.column > len(normalized_headers):
                    cell.alignment = styles['body_alignment']
                elif cell.fill is None or cell.fill == PatternFill():
                    cell.alignment = styles['body_alignment']

    _apply_sheet_styling(ws_rejected, styles)

    rejected_stream = io.BytesIO()
    wb_rejected.save(rejected_stream)
    rejected_stream.seek(0)
    return rejected_stream.getvalue()


def generate_validation_summary_workbook(rows, results, filename_base: str = 'Validation') -> bytes:
    """Generate a summary workbook with banking-style styling."""
    wb_summary = openpyxl.Workbook()
    ws_summary = wb_summary.active
    ws_summary.title = 'ValidationSummary'

    styles = _build_excel_styles()
    passed_count = sum(1 for result in results if result.get('status') == 'PASS')
    rejected_count = len(results) - passed_count
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')

    ws_summary.append(['Metric', 'Value'])
    ws_summary.append(['Total Records', len(rows)])
    ws_summary.append(['Passed', passed_count])
    ws_summary.append(['Rejected', rejected_count])
    ws_summary.append(['Timestamp', timestamp])

    for row_idx, row in enumerate(ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row), start=1):
        for cell in row:
            if row_idx == 1:
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
                cell.alignment = styles['header_alignment']
            else:
                cell.alignment = styles['body_alignment']

    ws_summary.cell(row=2, column=2).fill = styles['summary_pass_fill']
    ws_summary.cell(row=2, column=2).font = styles['status_valid_font']
    ws_summary.cell(row=3, column=2).fill = styles['summary_fail_fill']
    ws_summary.cell(row=3, column=2).font = styles['status_invalid_font']
    ws_summary.cell(row=4, column=2).fill = styles['summary_time_fill']
    ws_summary.cell(row=4, column=2).font = Font(color='FF4B5563', bold=True)

    _apply_sheet_styling(ws_summary, styles)

    summary_stream = io.BytesIO()
    wb_summary.save(summary_stream)
    summary_stream.seek(0)
    return summary_stream.getvalue()

    # Serialize to bytes
    passed_stream = io.BytesIO()
    wb_passed.save(passed_stream)
    passed_stream.seek(0)

    rejected_stream = io.BytesIO()
    wb_rejected.save(rejected_stream)
    rejected_stream.seek(0)

    return passed_stream.getvalue(), rejected_stream.getvalue()

@app.route('/health', methods=['GET'])
def health():
    return jsonify({"status": "ok"})

@app.route('/api/health', methods=['GET'])
def api_health():
    return health()

@app.route('/analyze', methods=['POST'])
def analyze_endpoint():
    print('[UPLOAD] Received upload')
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    uploaded_file = request.files['file']
    if uploaded_file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400

    try:
        file_bytes = uploaded_file.read()
    except Exception:
        return jsonify({"success": False, "message": "Uploaded file could not be read"}), 400

    if not file_bytes:
        return jsonify({"success": False, "message": "Uploaded file is empty"}), 400

    filename = uploaded_file.filename or ''
    mime_type = uploaded_file.mimetype or ''
    extension = Path(filename).suffix.lower()
    print(f"[UPLOAD] Extension: {extension} | MIME: {mime_type} | Size: {len(file_bytes)}")

    try:
        headers, rows, sheet_name = parse_excel_workbook(file_bytes, filename, mime_type)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        print(f"[UPLOAD] Excel read failed: {exc}")
        return jsonify({"success": False, "message": "Excel format unsupported"}), 400

    detected_profile = detect_profile(headers)
    print(f"[UPLOAD] Workbook loaded: sheet={sheet_name}, rows={len(rows)}")
    return jsonify({
        "success": True,
        "filename": uploaded_file.filename,
        "row_count": len(rows),
        "sheet_name": sheet_name,
        "profile": detected_profile
    })

@app.route('/validate', methods=['POST'])
def validate_endpoint():
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file part in request"}), 400

    uploaded_file = request.files['file']
    if uploaded_file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400

    selected_profile = request.form.get('profile', 'auto')

    # Read rules dynamically
    global RULES
    RULES = load_rules()

    try:
        file_bytes = uploaded_file.read()
    except Exception:
        return jsonify({"success": False, "message": "Uploaded file could not be read"}), 400

    if not file_bytes:
        return jsonify({"success": False, "message": "Uploaded file is empty"}), 400

    filename = uploaded_file.filename or ''
    mime_type = uploaded_file.mimetype or ''
    print(f"[VALIDATION] Received upload: filename={filename}, mime={mime_type}, size={len(file_bytes)}")

    try:
        headers, rows, _ = parse_excel_workbook(file_bytes, filename, mime_type)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        print(f"[VALIDATION] Excel read failed: {exc}")
        return jsonify({"success": False, "message": "Excel format unsupported"}), 400

    print(f"[VALIDATION] Workbook loaded: rows={len(rows)}")

    # Profile Detection
    detected_profile = detect_profile(headers)
    active_profile = selected_profile if selected_profile != 'auto' else detected_profile

    if active_profile == 'unknown':
        return jsonify({
            "error": "Could not auto-detect file profile. Please select profile manually from the dropdown."
        }), 400

    # Run validations
    start_time = time.time()
    results = validate_excel_data(rows, headers, active_profile, RULES)
    end_time = time.time()
    duration = f"{end_time - start_time:.2f} sec"

    # Extract original filename without extension, and get file extension
    original_filename = uploaded_file.filename or 'ValidationReport'
    file_name_base = Path(original_filename).stem  # Name without extension
    file_extension = Path(original_filename).suffix or '.xlsx'  # Keep original extension or default to .xlsx

    # Generate the passed and rejected validation reports workbooks with error details
    passed_report, rejected_report = generate_validated_excel_streams(rows, results, headers, active_profile)
    validation_summary = generate_validation_summary_workbook(rows, results, file_name_base)

    # Cache the resulting byte streams with original filename
    passed_count = sum(1 for r in results if r['status'] == 'PASS')
    rejected_count = sum(1 for r in results if r['status'] == 'FAIL')
    file_id = str(uuid.uuid4())
    VALIDATED_FILES_CACHE[file_id] = {
        "passed_report": passed_report,
        "rejected_report": rejected_report,
        "summary": validation_summary,
        "filename_base": file_name_base,
        "file_extension": file_extension,
        "passed_count": passed_count,
        "rejected_count": rejected_count,
        "total_count": len(rows),
        "full_results": results  # Store full results, not just first 20
    }

    # Generate summary lists
    error_breakdown = {
        "Invalid IBAN": 0,
        "Invalid Mobile": 0,
        "Invalid CNIC": 0,
        "Duplicate Records": 0,
        "Missing Required Fields": 0
    }

    for res in results:
        for err in res['errors']:
            msg = err['msg']
            field = err['field']
            if field == 'BeneficiaryIBAN':
                error_breakdown["Invalid IBAN"] += 1
            elif field in ['BeneficiaryMobile', 'BeneficiaryNumber']:
                error_breakdown["Invalid Mobile"] += 1
            elif field == 'BeneficiaryIdentificationNo' and 'CNIC' in msg:
                error_breakdown["Invalid CNIC"] += 1
            elif err.get('is_duplicate'):
                error_breakdown["Duplicate Records"] += 1
            elif 'is required' in msg or 'missing' in msg:
                error_breakdown["Missing Required Fields"] += 1

    return jsonify({
        "file_id": file_id,
        "filename": original_filename,
        "filename_base": file_name_base,
        "file_extension": file_extension,
        "profile": active_profile,
        "headers": headers,
        "rows": rows[:20],
        "results": results[:20],
        "duration": duration,
        "summary": {
            "total": len(rows),
            "passed": VALIDATED_FILES_CACHE[file_id]["passed_count"],
            "rejected": VALIDATED_FILES_CACHE[file_id]["rejected_count"],
            "errors_breakdown": error_breakdown
        }
    })

@app.route('/api/analyze', methods=['POST'])
def analyze_api_endpoint():
    return analyze_endpoint()

@app.route('/api/validate', methods=['POST'])
def validate_api_endpoint():
    return validate_endpoint()

@app.route('/download/report/<file_id>', methods=['GET'])
def download_report(file_id):
    if file_id not in VALIDATED_FILES_CACHE:
        return "File not found or expired", 404

    data = VALIDATED_FILES_CACHE[file_id]
    report_bytes = data.get("rejected_report")
    if not report_bytes:
        return "Validation report not available", 404

    return send_file(
        io.BytesIO(report_bytes),
        as_attachment=True,
        download_name='rejected_records.xlsx',
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/download/report/<file_id>', methods=['GET'])
def download_report_api(file_id):
    return download_report(file_id)

@app.route('/download/passed/<file_id>', methods=['GET'])
def download_passed(file_id):
    if file_id not in VALIDATED_FILES_CACHE:
        return "File not found or expired", 404

    data = VALIDATED_FILES_CACHE[file_id]
    report_bytes = data.get("passed_report")
    if not report_bytes:
        return "Passed records report not available", 404

    return send_file(
        io.BytesIO(report_bytes),
        as_attachment=True,
        download_name='passed_records.xlsx',
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/download/passed/<file_id>', methods=['GET'])
def download_passed_api(file_id):
    return download_passed(file_id)

@app.route('/download/rejected/<file_id>', methods=['GET'])
def download_rejected(file_id):
    if file_id not in VALIDATED_FILES_CACHE:
        return "File not found or expired", 404

    data = VALIDATED_FILES_CACHE[file_id]
    report_bytes = data.get("rejected_report")
    if not report_bytes:
        return "Rejected records report not available", 404

    return send_file(
        io.BytesIO(report_bytes),
        as_attachment=True,
        download_name='rejected_records.xlsx',
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/download/rejected/<file_id>', methods=['GET'])
def download_rejected_api(file_id):
    return download_rejected(file_id)

if __name__ == '__main__':
    print('Backend module executed directly. Use a WSGI or serverless runtime to host this app.')
    ensure_runtime_directories()
    print('Upload directory ready')
    print('Generated directory ready')
    print('Logs directory ready')
