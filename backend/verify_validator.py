import json
from pathlib import Path

import openpyxl

from app import validate_rows


ROOT = Path(__file__).resolve().parent
RULES = json.loads((ROOT / 'rules.json').read_text(encoding='utf-8'))


def assert_text_preserved(values):
    wb = openpyxl.Workbook()
    ws = wb.active
    for idx, value in enumerate(values, start=1):
        ws.cell(row=1, column=idx, value=value)
        ws.cell(row=1, column=idx).number_format = '@'
    wb.save(ROOT / 'tmp_text_format_check.xlsx')
    reloaded = openpyxl.load_workbook(ROOT / 'tmp_text_format_check.xlsx', data_only=True)
    ws2 = reloaded.active
    for idx, value in enumerate(values, start=1):
        loaded_value = ws2.cell(row=1, column=idx).value
        assert loaded_value == value, f'Expected {value!r} to be preserved, got {loaded_value!r}'
        assert not isinstance(loaded_value, float), f'Expected string for {value!r}, got float'
    try:
        (ROOT / 'tmp_text_format_check.xlsx').unlink(missing_ok=True)
    except Exception:
        # On Windows the file may be locked by another process; ignore cleanup errors
        pass


def run_verification():
    assert_text_preserved(['03112753114', '000123456', '4220140603751'])
    cases = [
        (
            'raast',
            [
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456702',
                    'BeneficiaryName': 'Ali Khan',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'CNIC',
                    'BeneficiaryIdentificationNo': '4220140603751',
                    'ProductTypeCode': 'P1',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456703',
                    'BeneficiaryName': 'Bad CNIC',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'CNIC',
                    'BeneficiaryIdentificationNo': '12345',
                    'ProductTypeCode': '',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456704',
                    'BeneficiaryName': 'Bad CNIC Format',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'CNIC',
                    'BeneficiaryIdentificationNo': 'A220140603751',
                    'ProductTypeCode': '',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456705',
                    'BeneficiaryName': 'NTN Valid',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'NTN',
                    'BeneficiaryIdentificationNo': '41640934',
                    'ProductTypeCode': '',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456706',
                    'BeneficiaryName': 'NTN Alphanumeric',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'NTN',
                    'BeneficiaryIdentificationNo': 'C1234567',
                    'ProductTypeCode': '',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456707',
                    'BeneficiaryName': 'NTN Bad Length',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'NTN',
                    'BeneficiaryIdentificationNo': '4164093',
                    'ProductTypeCode': '',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'PK36SCBL0000001123456708',
                    'BeneficiaryName': 'IBAN Valid',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'CNIC',
                    'BeneficiaryIdentificationNo': '4220140603754',
                    'ProductTypeCode': '',
                },
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryIBAN': 'pk36scbl0000001123456709',
                    'BeneficiaryName': 'IBAN Invalid',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                    'BeneficiaryIdentificationType': 'CNIC',
                    'BeneficiaryIdentificationNo': '4220140603755',
                    'ProductTypeCode': '',
                },
            ],
            [
                ('PASS', []),
                ('FAIL', ['CNIC must be exactly 13 digits']),
                ('FAIL', ['CNIC must be numeric']),
                ('PASS', []),
                ('PASS', []),
                ('FAIL', ['NTN must be exactly 8 characters']),
                ('PASS', []),
                ('FAIL', ['IBAN must be uppercase and 24 characters long']),
            ],
        ),
        (
            'web_based',
            [
                {
                    'Amount': '100.50',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryAccountNo': '000123456',
                    'BeneficiaryName': 'Ali Khan',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                },
                {
                    'Amount': '',
                    'BeneficiaryBankCode': '001',
                    'BeneficiaryAccountNo': '000123457',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryMobile': '03112753114',
                },
            ],
            [
                ('PASS', []),
                ('FAIL', ['BeneficiaryName is required', 'Amount is required']),
            ],
        ),
        (
            '1link',
            [
                {
                    'Amount': '100.50',
                    'BankName': 'Test Bank',
                    'BeneficiaryAccountNumber': '000123456',
                    'BeneficiaryName': 'Sadiq Ali',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryNumber': '03112753114',
                },
                {
                    'Amount': '',
                    'BankName': 'Test Bank',
                    'BeneficiaryAccountNumber': '000123457',
                    'BeneficiaryName': '',
                    'BeneficiaryCode': 'B1',
                    'ReferenceField1': 'R1',
                    'ReferenceField2': 'R2',
                    'BeneficiaryEmail': 'ali@example.com',
                    'BeneficiaryNumber': '03112753114',
                },
            ],
            [
                ('PASS', []),
                ('FAIL', ['Amount is required', 'BeneficiaryName is required']),
            ],
        ),
    ]

    for profile, rows, expected in cases:
        results = validate_rows(rows, profile, RULES, debug=True)
        for idx, (status, expected_messages) in enumerate(expected):
            result = results[idx]
            actual_status = result['status']
            actual_messages = [err['msg'] for err in result['errors']]
            if actual_status != status:
                raise AssertionError(f'[{profile}] row {idx + 1}: expected status {status} but got {actual_status}')
            for message in expected_messages:
                if message not in actual_messages:
                    raise AssertionError(f'[{profile}] row {idx + 1}: expected message {message!r} in {actual_messages}')
        print(f"[VERIFY] {profile}: PASSED")


if __name__ == '__main__':
    run_verification()
