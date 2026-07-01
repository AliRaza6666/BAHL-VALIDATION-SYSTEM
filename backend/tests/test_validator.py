import io
import json
import sys
from pathlib import Path

import openpyxl
import xlwt

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app import detect_excel_format, parse_excel_workbook, validate_rows, generate_validated_excel_streams, generate_highlighted_validation_report


def load_rules():
    with open(Path(__file__).resolve().parents[1] / 'rules.json', 'r', encoding='utf-8') as f:
        return json.load(f)


def test_raast_profile_edge_cases():
    rules = load_rules()
    rows = [
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456702',
            'BeneficiaryName': 'Ali Khan',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'CNIC',
            'BeneficiaryIdentificationNo': '4220140603751',
            'ProductTypeCode': 'P1',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456703',
            'BeneficiaryName': 'Bad CNIC',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'CNIC',
            'BeneficiaryIdentificationNo': '12345',
            'ProductTypeCode': '',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456704',
            'BeneficiaryName': 'Bad CNIC Format',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'CNIC',
            'BeneficiaryIdentificationNo': 'A220140603751',
            'ProductTypeCode': '',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456705',
            'BeneficiaryName': 'NTN Valid',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'NTN',
            'BeneficiaryIdentificationNo': '41640934',
            'ProductTypeCode': '',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456706',
            'BeneficiaryName': 'NTN Alphanumeric',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'NTN',
            'BeneficiaryIdentificationNo': 'C1234567',
            'ProductTypeCode': '',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456707',
            'BeneficiaryName': 'NTN Bad Length',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'NTN',
            'BeneficiaryIdentificationNo': '4164093',
            'ProductTypeCode': '',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'PK36SCBL0000001123456708',
            'BeneficiaryName': 'IBAN Valid',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'CNIC',
            'BeneficiaryIdentificationNo': '4220140603754',
            'ProductTypeCode': '',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryIBAN': 'pk36scbl0000001123456709',
            'BeneficiaryName': 'IBAN Invalid',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
            'BeneficiaryIdentificationType': 'CNIC',
            'BeneficiaryIdentificationNo': '4220140603755',
            'ProductTypeCode': '',
        },
    ]

    results = validate_rows(rows, 'raast', rules)

    assert results[0]['status'] == 'PASS'
    assert results[1]['status'] == 'FAIL'
    assert any('CNIC must be exactly 13 digits' in err['msg'] for err in results[1]['errors'])
    assert results[2]['status'] == 'FAIL'
    assert any('CNIC must be numeric' in err['msg'] for err in results[2]['errors'])
    assert results[3]['status'] == 'PASS'
    assert results[4]['status'] == 'PASS'
    assert results[5]['status'] == 'FAIL'
    assert any('NTN must be exactly 8 characters' in err['msg'] for err in results[5]['errors'])
    assert results[6]['status'] == 'PASS'
    assert results[7]['status'] == 'FAIL'
    assert any('IBAN must be uppercase and 24 characters long' in err['msg'] for err in results[7]['errors'])


def test_detect_excel_format_supports_xlsx_and_xls_bytes():
    xlsx_buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['Amount', 'BeneficiaryName'])
    ws.append(['100', 'Ali'])
    wb.save(xlsx_buffer)
    xlsx_bytes = xlsx_buffer.getvalue()

    xls_buffer = io.BytesIO()
    book = xlwt.Workbook()
    sheet = book.add_sheet('Sheet1')
    sheet.write(0, 0, 'Amount')
    sheet.write(0, 1, 'BeneficiaryName')
    sheet.write(1, 0, 100)
    sheet.write(1, 1, 'Ali')
    book.save(xls_buffer)
    xls_bytes = xls_buffer.getvalue()

    xlsx_result = detect_excel_format('sample.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', xlsx_bytes)
    assert xlsx_result['kind'] == 'xlsx'

    xls_result = detect_excel_format('sample.xls', 'application/vnd.ms-excel', xls_bytes)
    assert xls_result['kind'] == 'xls'

    invalid_result = detect_excel_format('sample.csv', 'text/csv', b'Amount,BeneficiaryName')
    assert invalid_result['kind'] == 'unsupported'


def test_parse_excel_workbook_reads_both_formats():
    xlsx_buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Amount', 'BeneficiaryName'])
    ws.append(['100', 'Ali'])
    wb.save(xlsx_buffer)

    headers, rows, sheet_name = parse_excel_workbook(xlsx_buffer.getvalue(), 'sample.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    assert headers == ['Amount', 'BeneficiaryName']
    assert rows[0]['Amount'] == '100'
    assert sheet_name == 'Sheet1'

    xls_buffer = io.BytesIO()
    book = xlwt.Workbook()
    sheet = book.add_sheet('Sheet1')
    sheet.write(0, 0, 'Amount')
    sheet.write(0, 1, 'BeneficiaryName')
    sheet.write(1, 0, 100)
    sheet.write(1, 1, 'Ali')
    book.save(xls_buffer)

    headers, rows, sheet_name = parse_excel_workbook(xls_buffer.getvalue(), 'sample.xls', 'application/vnd.ms-excel')
    assert headers == ['Amount', 'BeneficiaryName']
    assert rows[0]['Amount'] == '100'
    assert sheet_name == 'Sheet1'


def test_parse_excel_workbook_rejects_corrupt_xls_with_clean_message():
    try:
        parse_excel_workbook(b'not-an-excel', 'sample.xls', 'application/vnd.ms-excel')
    except ValueError as exc:
        assert str(exc) == 'Uploaded Excel file is corrupted.'
    else:
        raise AssertionError('Expected ValueError for corrupt xls input')


def test_generate_validated_excel_streams_emits_xlsx_for_xls_input():
    xls_buffer = io.BytesIO()
    book = xlwt.Workbook()
    sheet = book.add_sheet('Sheet1')
    sheet.write(0, 0, 'Amount')
    sheet.write(0, 1, 'BeneficiaryName')
    sheet.write(1, 0, 100)
    sheet.write(1, 1, 'Ali')
    book.save(xls_buffer)

    headers, rows, _ = parse_excel_workbook(xls_buffer.getvalue(), 'sample.xls', 'application/vnd.ms-excel')
    results = [
        {'status': 'PASS', 'errors': []},
        {'status': 'FAIL', 'errors': [{'col': 0, 'field': 'Amount', 'msg': 'Amount is required', 'expected': 'Non-empty value', 'actual': '(Empty)'}]}
    ]

    from app import generate_validated_excel_streams
    passed_bytes, rejected_bytes = generate_validated_excel_streams(rows, results, headers, 'raast')

    assert passed_bytes.startswith(b'PK')
    assert rejected_bytes.startswith(b'PK')

    passed_wb = openpyxl.load_workbook(io.BytesIO(passed_bytes), data_only=True)
    rejected_wb = openpyxl.load_workbook(io.BytesIO(rejected_bytes), data_only=True)

    assert passed_wb.sheetnames == ['Sheet']
    assert rejected_wb.sheetnames == ['Sheet']


def test_web_profile_allows_duplicate_beneficiary_account_numbers():
    rules = load_rules()
    rows = [
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryAccountNo': '000123456',
            'BeneficiaryName': 'Ali Khan',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
        },
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryAccountNo': '000123456',
            'BeneficiaryName': 'Ahmed Khan',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
        },
    ]

    results = validate_rows(rows, 'web_based', rules)

    assert results[0]['status'] == 'PASS'
    assert results[1]['status'] == 'PASS'
    assert all(not any(err.get('is_duplicate') for err in result['errors']) for result in results)


def test_generate_validated_excel_streams_adds_status_and_error_columns_with_styling():
    rows = [
        {'BeneficiaryName': 'Ali', 'BeneficiaryMobile': '03001234567'},
        {'BeneficiaryName': 'Ahmed', 'BeneficiaryMobile': '0345'},
    ]
    results = [
        {'status': 'PASS', 'errors': []},
        {'status': 'FAIL', 'errors': [
            {'col': 1, 'field': 'BeneficiaryMobile', 'msg': 'Mobile number format invalid', 'expected': '03XXXXXXXXX', 'actual': '0345'}
        ]},
    ]

    passed_bytes, rejected_bytes = generate_validated_excel_streams(rows, results, ['BeneficiaryName', 'BeneficiaryMobile'], 'web_based')

    passed_wb = openpyxl.load_workbook(io.BytesIO(passed_bytes), data_only=False)
    passed_ws = passed_wb.active
    assert passed_ws.cell(row=1, column=3).value == 'Validation Status'
    assert passed_ws.cell(row=2, column=3).value == 'VALID'
    assert passed_ws.cell(row=1, column=3).fill.fgColor.rgb == 'FF0E8348'

    rejected_wb = openpyxl.load_workbook(io.BytesIO(rejected_bytes), data_only=False)
    rejected_ws = rejected_wb.active
    assert rejected_ws.cell(row=1, column=len(['BeneficiaryName', 'BeneficiaryMobile']) + 1).value == 'Validation Status'
    assert rejected_ws.cell(row=1, column=len(['BeneficiaryName', 'BeneficiaryMobile']) + 2).value == 'Error Reason'
    assert rejected_ws.cell(row=1, column=len(['BeneficiaryName', 'BeneficiaryMobile']) + 3).value == 'Suggested Fix'
    assert rejected_ws.cell(row=2, column=2).fill.fgColor.rgb == 'FFFDE2E2'
    assert rejected_ws.cell(row=2, column=len(['BeneficiaryName', 'BeneficiaryMobile']) + 2).value is not None
    assert 'Mobile number format invalid' in str(rejected_ws.cell(row=2, column=len(['BeneficiaryName', 'BeneficiaryMobile']) + 2).value)


def test_generate_highlighted_validation_report_includes_all_rows_and_comments():
    rows = [
        {'Amount': '100', 'BeneficiaryName': 'Ali', 'BeneficiaryAccountNo': '123'},
        {'Amount': '', 'BeneficiaryName': '', 'BeneficiaryAccountNo': '123'},
    ]
    results = [
        {'status': 'PASS', 'errors': []},
        {'status': 'FAIL', 'errors': [
            {'col': 0, 'field': 'Amount', 'msg': 'Amount is required', 'expected': 'Non-empty value', 'actual': '(Empty)'},
            {'col': 1, 'field': 'BeneficiaryName', 'msg': 'BeneficiaryName is required', 'expected': 'Non-empty value', 'actual': '(Empty)'}
        ]},
    ]

    report_bytes = generate_highlighted_validation_report(rows, results, ['Amount', 'BeneficiaryName', 'BeneficiaryAccountNo'], 'web_based')
    assert report_bytes.startswith(b'PK')

    wb = openpyxl.load_workbook(io.BytesIO(report_bytes), data_only=True)
    assert wb.sheetnames == ['Validation Report']
    ws = wb.active
    assert ws.max_row == 3
    assert ws.cell(row=2, column=1).value == '100'
    assert ws.cell(row=3, column=1).value in (None, '')
    assert ws.cell(row=3, column=1).comment is not None
    assert 'Amount is required' in ws.cell(row=3, column=1).comment.text


def test_web_profile_bypasses_identity_and_enforces_mandatory_fields():
    rules = load_rules()
    rows = [
        {
            'Amount': '100.50',
            'BeneficiaryBankCode': '001',
            'BeneficiaryAccountNo': '000123456',
            'BeneficiaryName': 'Ali Khan',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
        },
        {
            'Amount': '',
            'BeneficiaryBankCode': '001',
            'BeneficiaryAccountNo': '000123457',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryMobile': '03112753114',
        },
    ]

    results = validate_rows(rows, 'web_based', rules)

    assert results[0]['status'] == 'PASS'
    assert results[1]['status'] == 'FAIL'
    assert any(err['field'] == 'BeneficiaryName' for err in results[1]['errors'])
    assert any(err['field'] == 'Amount' for err in results[1]['errors'])


def test_1link_profile_maps_mobile_and_account_fields():
    rules = load_rules()
    rows = [
        {
            'Amount': '100.50',
            'BankName': 'Test Bank',
            'BeneficiaryAccountNumber': '000123456',
            'BeneficiaryName': 'Sadiq Ali',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryNumber': '03112753114',
        },
        {
            'Amount': '',
            'BankName': 'Test Bank',
            'BeneficiaryAccountNumber': '000123457',
            'BeneficiaryName': '',
            'BeneficiaryCode': 'B1',
            'ReferenceField1': 'R1',
            'ReferenceField2': 'R2',
            'BeneficiaryEmail': 'ali@example.com',
            'BeneficiaryNumber': '03112753114',
        },
    ]

    results = validate_rows(rows, '1link', rules)

    assert results[0]['status'] == 'PASS'
    assert results[1]['status'] == 'FAIL'
    assert any(err['field'] == 'Amount' for err in results[1]['errors'])
    assert any(err['field'] == 'BeneficiaryName' for err in results[1]['errors'])
